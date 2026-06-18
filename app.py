#!/usr/bin/env python
# coding: utf-8

from __future__ import annotations

import json
import os
import re
import unicodedata
from datetime import datetime
from functools import lru_cache
from io import BytesIO
from pathlib import Path

import pandas as pd
from flask import Flask, render_template, request, send_file, redirect, url_for, session
from openpyxl.styles import Alignment, Font, PatternFill

# ─── CONFIG ──────────────────────────────────────────────────────────────────

BASE_DIR      = Path(__file__).resolve().parent
PARQUET_DIR   = Path(os.getenv("PARQUET_DIR",   BASE_DIR / "storage_parquet"))
DATA_DIR      = PARQUET_DIR / "DATA_PARQUET"
META_DIR      = PARQUET_DIR / "META_PARQUET"
TEMPLATES_DIR = Path(os.getenv("TEMPLATES_DIR", BASE_DIR / "templates"))
MAX_POINTS_PER_TRACE = int(os.getenv("MAX_POINTS_PER_TRACE", "800"))

print(f"BASE_DIR     : {BASE_DIR}  → existe: {BASE_DIR.exists()}")
print(f"PARQUET_DIR  : {PARQUET_DIR}  → existe: {PARQUET_DIR.exists()}")
print(f"DATA_DIR     : {DATA_DIR}  → existe: {DATA_DIR.exists()}")
print(f"META_DIR     : {META_DIR}  → existe: {META_DIR.exists()}")
print(f"TEMPLATES_DIR: {TEMPLATES_DIR}  → existe: {TEMPLATES_DIR.exists()}")
if DATA_DIR.exists():
    print(f"  → {len(list(DATA_DIR.glob('*.parquet')))} fichiers .parquet détectés")

MONTHS_FR = {
    "janv.": 1, "févr.": 2, "fevr.": 2, "mars": 3, "avr.": 4,
    "mai": 5, "juin": 6, "juil.": 7, "août": 8, "aout": 8,
    "sept.": 9, "oct.": 10, "nov.": 11, "déc.": 12, "dec.": 12,
    "jan": 1, "jan.": 1, "feb": 2, "feb.": 2, "mar": 3, "mar.": 3,
    "apr": 4, "apr.": 4, "may": 5, "jun": 6, "jun.": 6,
    "jul": 7, "jul.": 7, "aug": 8, "aug.": 8, "sep": 9, "sep.": 9,
    "oct": 10, "oct.": 10, "nov": 11, "nov.": 11, "dec": 12, "dec.": 12,
}

_EN_TO_FR = {
    "jan": "janv.", "feb": "févr.", "mar": "mars", "apr": "avr.",
    "may": "mai",   "jun": "juin",  "jul": "juil.", "aug": "août",
    "sep": "sept.", "oct": "oct.",  "nov": "nov.",  "dec": "déc.",
}

def _normalize_time_label(label: str) -> str:
    parts = label.strip().split()
    if not parts:
        return label
    prefix = parts[0].lower().rstrip(".")
    if prefix in _EN_TO_FR:
        parts[0] = _EN_TO_FR[prefix]
    return " ".join(parts)


# ─── BLOC 2 : utilitaires ────────────────────────────────────────────────────

def _slug(value: str, maxlen: int = 25) -> str:
    """Convertit 'Money Supply M1' → 'money_supply_m1', tronque pour noms d'onglets."""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_").lower()
    return (text or "export")[:maxlen]


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Génère un nom d'onglet Excel valide (≤31 char, unique, sans caractères interdits)."""
    forbidden = r'[\\/?*\[\]:]'
    clean = re.sub(forbidden, "_", str(name))[:28]
    candidate = clean
    i = 1
    while candidate in used:
        candidate = f"{clean[:25]}_{i}"
        i += 1
    used.add(candidate)
    return candidate


def detect_frequency(time_values: pd.Series) -> str:
    """Détecte 'annual' | 'quarterly' | 'monthly' | 'unknown'."""
    sample = time_values.dropna().astype(str).head(20).tolist()
    if not sample:
        return "unknown"
    if any(re.match(r"^Q[1-4]\s+\d{4}$", s.strip(), re.IGNORECASE) for s in sample):
        return "quarterly"
    first_word = sample[0].strip().split()[0].lower() if sample[0].strip() else ""
    if first_word in MONTHS_FR:
        return "monthly"
    if all(re.match(r"^\d{4}(\.0)?$", s.strip()) for s in sample[:5]):
        return "annual"
    return "unknown"


def parse_time_label(label: str, freq: str) -> tuple[int, int]:
    """Convertit un label temporel en (année, période_dans_année)."""
    label = str(label).strip()
    if freq == "annual":
        return int(float(label)), 1
    if freq == "quarterly":
        m = re.match(r"^Q([1-4])\s+(\d{4})$", label, re.IGNORECASE)
        if m:
            return int(m.group(2)), int(m.group(1))
    if freq == "monthly":
        parts = label.split()
        if len(parts) == 2 and parts[0].lower() in MONTHS_FR:
            return int(parts[1]), MONTHS_FR[parts[0].lower()]
    raise ValueError(f"Label temporel non parsable : {label!r} (freq={freq})")


# ─── BLOC 3 : scan_storage — lit catalog.parquet ─────────────────────────────

FREQ_FR = {"annual": "Annuelle", "monthly": "Mensuelle",
           "quarterly": "Trimestrielle", "unknown": "Inconnue"}

# Mapping fréquences Datastream → clés internes
_FREQ_MAP = {
    "M": "monthly", "monthly": "monthly",
    "Q": "quarterly", "quarterly": "quarterly",
    "A": "annual", "annual": "annual",
}

@lru_cache(maxsize=1)
def scan_storage() -> dict[str, dict]:
    """
    Lit catalog.parquet (généré par migrate_to_parquet.py).
    Retourne un dict { nom_indicateur: info_dict }.

    Clé = colonne 'name' du catalog (nom long Datastream).
    Chemins fichiers reconstruits depuis DATA_DIR / META_DIR + stem du fichier source.

    lru_cache(1) : le catalog est statique entre deux migrations.
    Pour invalider manuellement : scan_storage.cache_clear()
    """
    catalog_path = PARQUET_DIR / "catalog.parquet"

    if not catalog_path.exists():
        raise FileNotFoundError(
            f"catalog.parquet introuvable dans {PARQUET_DIR}\n"
            "→ Lance migrate_to_parquet.py d'abord."
        )

    df_cat = pd.read_parquet(catalog_path, engine="pyarrow")

    result: dict[str, dict] = {}

    for row in df_cat.itertuples(index=False):
        rics          = json.loads(row.rics)
        ric_to_market = json.loads(row.ric_to_market)

        # Fréquence : normalise "M"/"Q"/"A" + valeurs textuelles
        freq = _FREQ_MAP.get(str(row.frequency).strip(), "unknown")

        # Reconstruction time_min / time_max en tuple (année, période)
        try:
            time_min = parse_time_label(str(row.time_min), freq)
            time_max = parse_time_label(str(row.time_max), freq)
        except Exception:
            time_min = time_max = (0, 0)

        # Nom de l'indicateur = colonne 'name' du catalog
        indicator_name = str(row.stem)

        result[indicator_name] = {
            "slug":           str(row.stem),
            # Chemin DATA : DATA_PARQUET/<stem>.parquet
            "file":           DATA_DIR / f"{row.stem}.parquet",
            # Chemin META : META_PARQUET/<stem>.parquet  (suffixe .meta conservé)
            "meta_file":      META_DIR / f"{row.stem}.meta.parquet",
            "countries":      sorted(rics, key=lambda ric: ric_to_market.get(ric, ric).lower()),
            "ric_to_country": ric_to_market,
            "frequency":      freq,
            "freq_fr":        FREQ_FR.get(freq, freq),
            "time_min":       time_min,
            "time_max":       time_max,
            # time_labels reconstruit depuis time_min/time_max (non stocké dans catalog)
            # → chargé à la demande dans load_indicator()
        }

    return result


# ─── BLOC 4 ──────────────────────────────────────────────────────────────────

def _extract_year(label) -> int:
    """Convertit n'importe quel label temporel en année int."""
    if isinstance(label, (int, float)):
        return int(label)
    if isinstance(label, pd.Timestamp):
        return label.year
    s = str(label).strip()
    try:
        return pd.to_datetime(s, errors="raise", dayfirst=False).year
    except Exception:
        pass
    m = re.search(r"(\d{4})", s)
    if m:
        return int(m.group(1))
    return 0


# ─── BLOC 5 ──────────────────────────────────────────────────────────────────

def filter_time_window(data: pd.DataFrame, freq: str,
                       year_start: int, year_end: int) -> pd.DataFrame:
    """Filtre les lignes dont l'année est dans [year_start, year_end]."""
    time_col = data.columns[0]
    years = data[time_col].astype(str).apply(lambda lbl: parse_time_label(lbl, freq)[0])
    return data.loc[years.between(year_start, year_end)].copy()


def convert_to_annual(data: pd.DataFrame, freq: str) -> pd.DataFrame:
    """
    Convertit en annuel via moyenne sur années calendaires complètes uniquement.
    Annual → retourne tel quel. Quarterly → 4 trimestres requis. Monthly → 12 mois requis.
    """
    if freq == "annual":
        return data

    time_col   = data.columns[0]
    parsed     = data[time_col].astype(str).apply(lambda lbl: parse_time_label(lbl, freq))
    df         = data.copy()
    df["__year"] = [p[0] for p in parsed]

    required   = {"quarterly": 4, "monthly": 12}.get(freq, 1)
    value_cols = [c for c in df.columns if c not in (time_col, "__year")]

    rows = []
    for year, sub in df.groupby("__year"):
        if sub[time_col].nunique() < required:
            continue
        row = {"Année": year}
        for col in value_cols:
            row[col] = sub[col].mean(skipna=True)
        rows.append(row)

    return pd.DataFrame(rows)


# ─── BLOC 6 : export Excel ───────────────────────────────────────────────────

def build_excel_export(selections: list[dict], year_start: int, year_end: int) -> BytesIO:
    """
    Construit le .xlsx final.
    Onglet README + DATA_<slug> + META_<slug> par indicateur.
    Fréquence native conservée. Filtre temporel par année extraite.
    """
    buffer     = BytesIO()
    used_names: set[str] = {"README"}

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:

        readme_rows = [
            {"Champ": "Date d'export",      "Valeur": datetime.now().strftime("%Y-%m-%d %H:%M")},
            {"Champ": "Fenêtre temporelle", "Valeur": f"{year_start} – {year_end}"},
            {"Champ": "", "Valeur": ""},
            {"Champ": "── Indicateurs inclus ──", "Valeur": ""},
        ]
        for sel in selections:
            readme_rows.append({
                "Champ":  sel["indicator"],
                "Valeur": f"Fréquence : {FREQ_FR.get(sel['freq'], sel['freq'])} | "
                          f"{len(sel['countries'])} pays",
            })
        readme_rows += [
            {"Champ": "", "Valeur": ""},
            {"Champ": "── Note méthodologique ──", "Valeur": ""},
            {"Champ": "Fréquences hétérogènes",
             "Valeur": "Chaque indicateur conserve sa fréquence native. "
                       "Pour comparer des séries de fréquences différentes, "
                       "agréger manuellement (ex. moyenne annuelle) via TCD."},
            {"Champ": "Source", "Valeur": "LSEG Datastream via Workspace"},
        ]
        pd.DataFrame(readme_rows).to_excel(writer, sheet_name="README", index=False)

        for sel in selections:
            ind            = sel["indicator"]
            data           = sel["data"]
            meta           = sel["meta"]
            freq           = sel["freq"]
            countries      = sel["countries"]
            ric_to_country = sel.get("ric_to_country", {})

            time_col  = data.columns[0]
            keep_cols = [time_col] + [c for c in countries if c in data.columns]
            data      = data[keep_cols].copy()

            for c in countries:
                if c not in data.columns:
                    data[c] = pd.NA
            data = data[[time_col] + countries]

            rename_map = {ric: f"{ric_to_country.get(ric, ric)} ({ric})" for ric in countries}
            data = data.rename(columns=rename_map)

            years = data[time_col].apply(_extract_year)
            data  = data.loc[years.between(year_start, year_end)].copy()

            slug       = _slug(ind)
            data_sheet = _safe_sheet_name(f"DATA_{slug}", used_names)
            meta_sheet = _safe_sheet_name(f"META_{slug}", used_names)

            data.to_excel(writer, sheet_name=data_sheet, index=False)

            # Filtre META sur les RIC sélectionnés uniquement

def build_excel_export(payload, year_start, year_end):
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        used_names = set()
        
        for ind, countries in payload_dict.items():
            # ... tout ton code de traitement ...
            
            data.to_excel(writer, sheet_name=data_sheet, index=False)

            # ── META filtrée sur RIC sélectionnés ──────────────────────────
            if not meta.empty and "RIC" in meta.columns:
                meta_out = meta[meta["RIC"].astype(str).str.strip().isin(countries)].copy()
                if meta_out.empty:
                    meta_out = pd.DataFrame({"Champ": ["(aucun RIC correspondant en metadata)"]})
            else:
                meta_out = pd.DataFrame({"Champ": ["(metadata source absente)"]})

            extra = pd.DataFrame({
                "Champ":  ["Fenêtre temporelle appliquée", "Fréquence native", "Date export"],
                "Valeur": [f"{year_start} – {year_end}",
                           FREQ_FR.get(freq, freq),
                           datetime.now().strftime("%Y-%m-%d %H:%M")],
            })
            if not meta.empty and "RIC" in meta.columns:
                meta_out = pd.concat(
                    [meta_out, pd.DataFrame([{c: "" for c in meta_out.columns}]), extra],
                    ignore_index=True
                )
            else:
                meta_out = extra

            meta_out.to_excel(writer, sheet_name=meta_sheet, index=False)
            # ── fin boucle ──────────────────────────────────────────────────

        _style_workbook(writer)   # ← hors boucle, une seule fois

    buffer.seek(0)
    return buffer



def _style_workbook(writer: pd.ExcelWriter) -> None:
    """En-têtes gras + fond bleu + largeurs auto."""
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for sheet_name in writer.book.sheetnames:
        ws = writer.book[sheet_name]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")
        for column_cells in ws.columns:
            values = [str(c.value) if c.value is not None else "" for c in column_cells]
            width  = min(max((len(v) for v in values), default=10) + 2, 50)
            ws.column_dimensions[column_cells[0].column_letter].width = width


# ─── BLOC 6bis : chargement indicateur depuis Parquet ────────────────────────

@lru_cache(maxsize=64)
def _load_indicator_cached(stem: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Charge DATA + META depuis les Parquet.
    Cache LRU sur le stem (nom de fichier sans extension).
    Invalider avec _load_indicator_cached.cache_clear() après migration.

    Retourne (df_data, df_meta) — df_meta peut être vide.
    """
    data_path = DATA_DIR / f"{stem}.parquet"
    meta_path = META_DIR / f"{stem}.meta.parquet"

    df_data = pd.read_parquet(data_path, engine="pyarrow")

    try:
        df_meta = pd.read_parquet(meta_path, engine="pyarrow")
    except Exception:
        df_meta = pd.DataFrame()

    return df_data, df_meta


def load_indicator(indicator_name: str) -> tuple[pd.DataFrame, pd.DataFrame, str]:
    """
    Interface publique : charge un indicateur par son nom long.
    Retourne (data, meta, freq).
    """
    catalog = scan_storage()
    if indicator_name not in catalog:
        raise KeyError(f"Indicateur inconnu : {indicator_name!r}")

    info            = catalog[indicator_name]
    df_data, df_meta = _load_indicator_cached(info["slug"])

    return df_data, df_meta, info["frequency"]


# ─── BLOC 7 : Flask routes ───────────────────────────────────────────────────

app = Flask(__name__, template_folder=str(TEMPLATES_DIR))
app.secret_key = "medef-portail-dev-key-change-in-prod"


def _build_view(catalog: dict) -> tuple[dict, dict, dict]:
    """Prépare 3 dicts pour le template step1."""
    view          = {ind: info["countries"]              for ind, info in catalog.items()}
    country_names = {ind: info.get("ric_to_country", {}) for ind, info in catalog.items()}
    freq_labels   = {ind: info["freq_fr"]                for ind, info in catalog.items()}
    return view, country_names, freq_labels


@app.get("/")
def step1():
    try:
        catalog = scan_storage()
        view, country_names, freq_labels = _build_view(catalog)
        return render_template("step1.html",
            catalog=view, country_names=country_names,
            freq_labels=freq_labels, error=None)
    except Exception as exc:
        return render_template("step1.html",
            catalog={}, country_names={}, freq_labels={}, error=str(exc))


@app.post("/configure")
def step2():
    indicators = request.form.getlist("indicators")
    selections = {ind: [] for ind in indicators}
    for raw in request.form.getlist("country_pairs"):
        if "::" in raw:
            ind, pays = raw.split("::", 1)
            if ind in selections:
                selections[ind].append(pays)

    selections = {ind: pays for ind, pays in selections.items() if pays}
    if not selections:
        return redirect(url_for("step1"))

    session["selections"] = selections

    catalog       = scan_storage()
    country_names = {ind: catalog[ind].get("ric_to_country", {}) for ind in selections}
    freq_labels   = {ind: catalog[ind]["freq_fr"]                for ind in selections}

    def _year(v):
        return v[0] if isinstance(v, (tuple, list)) else v

    mins     = [_year(catalog[ind]["time_min"]) for ind in selections if catalog[ind].get("time_min")]
    maxs     = [_year(catalog[ind]["time_max"]) for ind in selections if catalog[ind].get("time_max")]
    year_min = min(mins) if mins else 1945
    year_max = max(maxs) if maxs else datetime.now().year

    charts_data = {}
    for ind, rics in selections.items():
        data, _, freq = load_indicator(ind)
        time_col      = data.columns[0]
        charts_data[ind] = {}
        for ric in rics:
            if ric not in data.columns:
                continue
            x_years = data[time_col].astype(str).apply(_extract_year).tolist()
            y_vals  = pd.to_numeric(data[ric], errors="coerce").where(
                          pd.to_numeric(data[ric], errors="coerce").notna(), other=None
                      ).tolist()
            charts_data[ind][ric] = {
                "x":    x_years,
                "y":    y_vals,
                "name": country_names[ind].get(ric, ric),
            }

    return render_template("step2.html",
        selections=selections,
        country_names=country_names,
        freq_by_indicator=freq_labels,
        year_min=year_min,
        year_max=year_max,
        charts_data=charts_data)


@app.post("/download")
def download():
    try:
        selections = session.get("selections", {})
        if not selections:
            return redirect(url_for("step1"))

        year_start = int(request.form.get("year_start"))
        year_end   = int(request.form.get("year_end"))
        if year_start > year_end:
            year_start, year_end = year_end, year_start

        catalog = scan_storage()
        payload = []
        for ind, countries in selections.items():
            if not countries:
                continue
            data, meta, freq = load_indicator(ind)
            payload.append({
                "indicator":      ind,
                "countries":      countries,
                "data":           data,
                "meta":           meta,
                "freq":           freq,
                "ric_to_country": catalog[ind].get("ric_to_country", {}),
            })

        buffer    = build_excel_export(payload, year_start, year_end)
        slug_inds = "_".join(_slug(p["indicator"], maxlen=15) for p in payload)[:80]
        filename  = f"export_{slug_inds}_{datetime.now().strftime('%Y%m%d')}.xlsx"

        return send_file(buffer, as_attachment=True, download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as exc:
        return f"Erreur lors de la génération : {exc}", 400


@app.get("/health")
def health():
    scan_storage()
    return {"status": "ok"}


print("[ROUTES]")
for rule in app.url_map.iter_rules():
    print(f"  {sorted(rule.methods)} {rule.rule}")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=False)
